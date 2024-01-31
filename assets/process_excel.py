import openpyxl
import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from mapping_list import mapping_list

def populate_boolean(ws):
    tooltype_col = 3
    last_row_with_data = max(row_num for row_num, cell_values in enumerate(ws.iter_rows(min_col=tooltype_col, max_col=tooltype_col, values_only=True), start=1) if cell_values[0] is not None)

    for row in range(2, last_row_with_data + 1):
        cell_c = ws.cell(column=3, row=row)
        if cell_c.value is None:
            continue

        for column in range(4, 9):
            cell = ws.cell(column=column, row=row)
            bgColor = cell.fill.bgColor.index
            if cell.value is None:
                cell.value = 1 if bgColor != '00000000' else 0

def process_formula(cell, column):
    formula_columns = [27, 28, 29, 30]
    if column in formula_columns and isinstance(cell.value, str) and '-' in cell.value:
        parts = cell.value.split('-')
        try:
            result = float(parts[0])
            for part in parts[1:]:
                result -= float(part)
            return result
        except (ValueError, IndexError):
            return cell.value
    return cell.value

def trim_empty_columns(df):
    for col in reversed(df.columns):
        if df[col].dropna().eq('emptyCell').all():
            df.drop(columns=col, inplace=True)
        else:
            break
    return df

def xls_to_csv(df):
    print("Converting xls to csv...")
    df.replace(to_replace=[r"\n", r"\s\s+"], value=[" ", " "], regex=True, inplace=True)
    df.fillna('emptyCell', inplace=True)
    trim_empty_columns(df)
    csv_path = "output_list.csv"
    df.to_csv(csv_path, sep=";", index=False, header=False)
    return csv_path

def clean_csv(csv_path):
    print("Cleaning csv...")
    df = pd.read_csv(csv_path, sep=";", header=None, dtype=str, na_values=[], keep_default_na=False)
    df = df[df[2] != "emptyCell"]
    df = df.apply(lambda col: col.str.strip())
    df = df[~df.apply(lambda row: row.str.isspace().all() or row.isnull().all(), axis=1)]
    df.drop(columns=[0, 1], inplace=True)
    output_csv_path = "output_list_modified.csv"
    df.to_csv(output_csv_path, sep=";", index=False, header=False)
    return output_csv_path

def separate_csv(csv_path):
    print("Separating csv files...")
    newHeader = 'tooltype;steel;stainless;castiron;aluminum;universal;catnum;invnum;unit;grinded;mfr;holdertype;tipdia;shankdia;pitch;neckdia;tslotdp;toollen;splen;worklen;bladecnt;tiptype;tipsize;material;coating;inserttype;cabinet;qty;issued;avail;minqty;secocab;sandvikcab;kennacab;niagaracab;extcab;sourcetable'
    
    df = pd.read_csv(csv_path, sep=";", header=None, dtype=str, na_values=[], keep_default_na=False)
    df.drop(index=0, inplace=True)

    df.insert(8, 'unit', df[7].apply(lambda x: 'inch' if x and x[1] == '2' else 'mm'))
    
    for col in df.columns:
        if col != 26:
            df[col] = df[col].apply(lambda x: x.replace(",", ".") if x != "N/A" else x)
            
    df.drop(df.tail(1).index, inplace=True)

    fixture_df = df[df[7].str.startswith(('A', 'C'), na=False)]
    thread_making_df = df[df[7].str.startswith(('119', '129', '411', '421', '412', '422', '511', '521'), na=False)]
    
    df = df[~df.index.isin(fixture_df.index)]
    df = df[~df.index.isin(thread_making_df.index)]

    df['sourcetable'] = 'tool'
    fixture_df['sourcetable'] = 'fixture'
    thread_making_df['sourcetable'] = 'threadmaking'

    tool_table_csv_path = "tool_table.csv"
    with open(tool_table_csv_path, 'w') as f:
        f.write(newHeader + '\n')
        df.to_csv(f, sep=";", index=False, header=False)

    if not fixture_df.empty:
        fixture_table_csv_path = "fixture_table.csv"
        with open(fixture_table_csv_path, 'w') as f:
            f.write(newHeader + '\n')
            fixture_df.to_csv(f, sep=";", index=False, header=False)

    if not thread_making_df.empty:
        thread_making_csv_path = "thread_making_table.csv"
        with open(thread_making_csv_path, 'w') as f:
            f.write(newHeader + '\n')
            thread_making_df.to_csv(f, sep=";", index=False, header=False)

    return tool_table_csv_path


def add_diameter_columns(csv_path):
    print("Adding diameter columns...")
    df = pd.read_csv(csv_path, sep=";", dtype=str, na_values=[], keep_default_na=False)
    
    df["tipdia_mm"] = ""
    df["tipdia_inch"] = ""
    
    for index, row in df.iterrows():
        try:
            tipdia = float(row["tipdia"])
            unit = row["unit"]
            
            if unit == "mm":
                df.at[index, "tipdia_mm"] = row["tipdia"]
                df.at[index, "tipdia_inch"] = "{:.4f}".format(tipdia / 25.4)
            elif unit == "inch":
                df.at[index, "tipdia_inch"] = row["tipdia"]
                df.at[index, "tipdia_mm"] = "{:.2f}".format(tipdia * 25.4)
        except ValueError:
            continue


    df.drop(columns=["tipdia"], inplace=True)
    
    holdertype_col_index = df.columns.get_loc("holdertype")

    cols = list(df.columns)
    df = df[cols[:holdertype_col_index + 1] + ["tipdia_mm", "tipdia_inch"] + cols[holdertype_col_index + 1:-2]]

    df.to_csv(csv_path, sep=";", index=False)

def add_diameter_columns_no_conversion(csv_path):
    df = pd.read_csv(csv_path, sep=";", dtype=str, na_values=[], keep_default_na=False)


    if 'tipdia' in df.columns:
        df["tipdia_mm"] = ""
        df["tipdia_inch"] = ""

        for index, row in df.iterrows():
            tipdia = row["tipdia"]
            unit = row["unit"]

            if unit == "mm":
                df.at[index, "tipdia_mm"] = tipdia
            elif unit == "inch":
                df.at[index, "tipdia_inch"] = tipdia

        holdertype_col_index = df.columns.get_loc("holdertype")

        cols = list(df.columns)
        df = df[cols[:holdertype_col_index + 1] + ["tipdia_mm", "tipdia_inch"] + cols[holdertype_col_index + 1:-2]]

        df.drop(columns=["tipdia"], inplace=True)

    df.to_csv(csv_path, sep=";", index=False)

def add_subtype_column(csv_path):
    print("Adding subtype columns...")
    df = pd.read_csv(csv_path, sep=";", dtype=str, na_values=[], keep_default_na=False)

    if 'tooltype' in df.columns:
        df['subtype'] = df['tooltype']

    df.to_csv(csv_path, sep=";", index=False)

mapping_dict = {item[1]: item[2] for item in mapping_list}

def update_tooltype_based_on_subtype(csv_path):
    print("Updating tooltypes...")
    df = pd.read_csv(csv_path, sep=";", dtype=str, na_values=[], keep_default_na=False)

    if 'subtype' in df.columns and 'tooltype' in df.columns:
        df['tooltype'] = df['subtype'].apply(lambda x: mapping_dict.get(x, x))

    df.to_csv(csv_path, sep=";", index=False)

def cleanup_files(*files_to_remove):
    for file in files_to_remove:
        if os.path.exists(file):
            os.remove(file)

def main_gui():
    def process_file():
        input_file = file_path_var.get()
        print("Opening file...")
        try:
            wb = openpyxl.load_workbook(input_file, data_only=True)
            fs = wb.active
        except Exception as e:
            result_var.set(f"Error loading the Excel file: {e}")
            return
        print("Populating booleans")
        populate_boolean(fs)
        print("Finished populating booleans")

        data = []
        for row_index, row in enumerate(fs.iter_rows(), 1):
            data_row = [process_formula(cell, col_idx) for col_idx, cell in enumerate(row, 1)]
            data.append(data_row)

        df_processed = pd.DataFrame(data)
        processed_csv = xls_to_csv(df_processed)
        cleaned_csv = clean_csv(processed_csv)
        final_csv = separate_csv(cleaned_csv)
        
        add_diameter_columns(final_csv)

        add_diameter_columns_no_conversion("fixture_table.csv")
        add_diameter_columns_no_conversion("thread_making_table.csv")

        add_subtype_column(final_csv)
        add_subtype_column("fixture_table.csv")
        add_subtype_column("thread_making_table.csv")

        update_tooltype_based_on_subtype(final_csv)
        update_tooltype_based_on_subtype("fixture_table.csv")
        update_tooltype_based_on_subtype("thread_making_table.csv")


        result_var.set(f'Final CSV ready at {final_csv}')
        cleanup_files(processed_csv, cleaned_csv)
        print("Finished!")


    def open_file():
        file_path = filedialog.askopenfilename(title="Open Excel File", filetypes=[("Excel Files", "*.xlsx")])
        file_path_var.set(file_path)

    root = tk.Tk()
    root.title("Excel to CSV Converter")

    file_path_var = tk.StringVar()
    result_var = tk.StringVar()

    frame = tk.Frame(root, padx=20, pady=20)
    frame.pack(padx=10, pady=10)

    open_btn = tk.Button(frame, text="Open Excel File", command=open_file)
    open_btn.pack(pady=10)

    path_label = tk.Label(frame, textvariable=file_path_var)
    path_label.pack(pady=10)

    process_btn = tk.Button(frame, text="Process File", command=process_file)
    process_btn.pack(pady=10)

    result_label = tk.Label(frame, textvariable=result_var)
    result_label.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    main_gui()